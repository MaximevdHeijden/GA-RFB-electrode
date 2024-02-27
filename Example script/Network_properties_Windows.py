import openpnm as op
import porespy as ps
import numpy as np
from skimage import io
from skimage import util
import openpyxl

# Load in project:   
path = '.\\Genetic_Algorithm\\GA-2\\'
name = 'j0_1_generation_0_individual_3'
project_1 = op.io.PNM.load_project(filename=path+name+'.pnm')
net = project_1.network
project_1 = op.io.PNM.load_project(filename=path+name+'.pnm')
net = project_1.network
Type = 'GA\\'

def find_connectivity(net):
    r''' find_connectivity finds the connectivity of every pore in a network.

    Parameters
    ----------
    net : The network for which you would like to find the connectivity.

    Returns
    -------
    unique : Amount of pores a pore is connected to.
    counts : The amount of pores which have this unique connectivity.
    (e.g. for unique = 1, counts = 3, 3 pores have a connectivity of 1)
    norm_counts : counts normalized to unity. 

    '''
    conn = np.zeros(len(net.pores('internal'))) # The connectivity array, displaying the amount of pores every pore is connected to.
    nr=0
    
    for i in net.pores('internal'):
        conn[nr] = len(net.find_neighbor_pores(pores=i))
        nr=nr+1
      
    unique, counts = np.unique(conn, return_counts=True) 
    norm_counts = counts/np.sum(counts)
    
    return unique, norm_counts, conn

def pore_size_distribution(bin_width, pore_volume, pore_diameter, pores):     
    r''' pore_size_distribution returns the normalized volume and cumulative
    normalized volume for a given set of pores, which can be used to plot the
    pore size distribution histogram.
    Parameters
    ----------
    bin_width : Specified bin_width for the histogram.
    pores : The pores that are analyzed.
    pore_volume : Array of pore volumes of every pore in pores.
    pore_diameter : Array of pore diameters of every pore in pores.
            
    Returns
    -------
    pore_size_range : The position of the left side of the bins.
    bins : The number of pores in every bin 
    norm_volume : The normalized volume that every bin accounts for.
    cum_volume : THe cumulative normalized volume that every bin up to that bin
    accounts for. 
    '''
    pore_diameter_2 = pore_diameter * 1e6 # Conversion from [m] to [um]
    norm_pore_volume_2 = pore_volume/np.sum(pore_volume) # Normalize pore volume to unity
    
    # Set up iterating variables
    max_pore_size = int(max(np.ceil(pore_diameter_2/bin_width)*bin_width))
    pore_size_range = range(0, max_pore_size, bin_width)
    norm_volume = np.zeros(len(pore_size_range)) 
    norm_volume_2 = np.zeros(len(pore_size_range)) 
    bins = np.zeros(len(pore_size_range))
    i = 0
    
    # Divide pores over bins
    for size_bin in pore_size_range:
        nr = 0
        for pore in pores:
            if size_bin < pore_diameter_2[nr] <= size_bin + bin_width:
                
                bins[i] += 1
                norm_volume[i] += 1 
                norm_volume_2[i] += norm_pore_volume_2[nr]
            nr = nr+1
        i += 1
        
    # Test if all pores are assigned    
    print(f'Total normalized volume (should be 1): {sum(norm_volume) :.2f}')
    
    # Compute cumulative volume
    j = 0
    cum_volume_iter = 0
    cum_volume = np.zeros(len(pore_size_range))
    
    for bin_volume in norm_volume_2:
        cum_volume[j] = cum_volume_iter + bin_volume
        cum_volume_iter += bin_volume
        j += 1
    
    psd = {}
    psd['pore_size_range'] = pore_size_range
    psd['bins'] = bins
    psd['norm_volume'] = norm_volume
    psd['cum_volume'] = cum_volume
    return psd

def throat_size_distribution(bin_width, throat_volume, throat_diameter, throats):     
    r''' pore_size_distribution returns the normalized volume and cumulative
    normalized volume for a given set of pores, which can be used to plot the
    pore size distribution histogram.
    Parameters
    ----------
    bin_width : Specified bin_width for the histogram.
    pores : The pores that are analyzed.
    pore_volume : Array of pore volumes of every pore in pores.
    pore_diameter : Array of pore diameters of every pore in pores.
            
    Returns
    -------
    pore_size_range : The position of the left side of the bins.
    bins : The number of pores in every bin 
    norm_volume : The normalized volume that every bin accounts for.
    cum_volume : THe cumulative normalized volume that every bin up to that bin
    accounts for. 
    '''
    throat_diameter_2 = throat_diameter * 1e6 # Conversion from [m] to [um]
    norm_throat_volume_2 = throat_volume/np.sum(throat_volume) # Normalize pore volume to unity
    
    # Set up iterating variables
    max_throat_size = int(max(np.ceil(throat_diameter_2/bin_width)*bin_width))
    throat_size_range = range(0, max_throat_size, bin_width)
    norm_volume = np.zeros(len(throat_size_range)) 
    norm_volume_2 = np.zeros(len(throat_size_range)) 
    bins = np.zeros(len(throat_size_range))
    i = 0
       
    # Divide pores over bins
    for size_bin in throat_size_range:
        nr = 0
        for throat in throats:
            if size_bin < throat_diameter_2[nr] <= size_bin + bin_width:
                
                bins[i] += 1
                norm_volume[i] += 1 
                norm_volume_2[i] += norm_throat_volume_2[nr]
            nr = nr+1
        i += 1
        
    # Test if all pores are assigned    
    print(f'Total normalized volume (should be 1): {sum(norm_volume) :.2f}')
    
    # Compute cumulative volume
    j = 0
    cum_volume_iter = 0
    cum_volume = np.zeros(len(throat_size_range))
    
    for bin_volume in norm_volume_2:
        cum_volume[j] = cum_volume_iter + bin_volume
        cum_volume_iter += bin_volume
        j += 1
    
    tsd = {}
    tsd['throat_size_range'] = throat_size_range
    tsd['bins'] = bins
    tsd['norm_volume'] = norm_volume
    tsd['cum_volume'] = cum_volume
    return tsd    

def find_porosity(net, im, voxel_size):
    r'''
    Finds the porosity calculated from the network and from the corresponding
    image.

    Parameters
    ----------
    net : Pore network of the electrode of interest.
    im : Corresponding image with the void space set as True.
    voxel_size : Voxel_size of the image in [meter].

    Returns
    -------
    net_porosity : Porosity calculated from the network.
    im_porisity : Porosity calculated from the image.
    '''
    
    pore_volume = net['pore.volume'][net.pores('internal')]  
    void_volume = pore_volume.sum()
    total_volume = im.shape[0]*im.shape[1]*im.shape[2]*voxel_size**3
    net_porosity = void_volume/total_volume
    im_porosity = ps.metrics.porosity(im)
    
    return net_porosity, im_porosity

def permeability(net, im, voxel_size):
    r'''
    Finds the permeability vector for the given network.
    
    NOTE: Check if net.pores('left, right, front, back, bottom and top') correspond
    with the right faces for inlet and outlet with the given image shape in Paraview
    before running! 
    
    In this case:   'left, right' = x direction
                    'front, back' = y direction
                    'top, bottom' = z direction
    
    Parameters
    ----------
    net : Pore network of the electrode of interest.
    im : Corresponding image with the void space set as True.
    voxel_size : Voxel_size of the image in [meter].

    Returns
    -------
    K: Permeability vector [Kx, Ky, Kz]
    '''
    
    print('NOTE: When running permeability(), make sure that the '
          'assigned pores match the dimensions (image shape) given in this function!')
    
    # Network flow cross-sectional area
    area = {
    'x': im.shape[1] * im.shape[2] * voxel_size**2,
    'y': im.shape[0] * im.shape[2] * voxel_size**2,
    'z': im.shape[0] * im.shape[1] * voxel_size**2, 
        }

    # Network flow length
    length = {
    'x': im.shape[0] * voxel_size,
    'y': im.shape[1] * voxel_size,
    'z': im.shape[2] * voxel_size,
        }
    
    # Add geometry
    geo = op.geometry.GenericGeometry(network=net, pores=net.Ps, throats=net.Ts)

    # Add phase and physics
    air = op.phases.Air(network=net)
    phys_air = op.physics.Standard(network=net, phase=air, geometry=geo)

    # Perform StokesFlow algorithm to calculate x Permeability
    sf_x = op.algorithms.StokesFlow(network=net)
    sf_x.setup(phase=air)
    sf_x.set_value_BC(pores=net.pores('left'), values=101325)
    sf_x.set_value_BC(pores=net.pores('right'), values=0)
    sf_x.run()
    air.update(sf_x.results())

    K_x = sf_x.calc_effective_permeability(domain_area=area['x'], domain_length=length['x'])[0]

    # Perform StokesFlow algorithm to calculate y Permeability
    sf_y = op.algorithms.StokesFlow(network=net)
    sf_y.setup(phase=air)
    sf_y.set_value_BC(pores=net.pores('front'), values=101325)
    sf_y.set_value_BC(pores=net.pores('back'), values=0)
    sf_y.run()
    air.update(sf_y.results())

    K_y = sf_y.calc_effective_permeability(domain_area=area['y'], domain_length=length['y'])[0]

    # Perform StokesFlow algorithm to calculate z Permeability
    sf_z = op.algorithms.StokesFlow(network=net)
    sf_z.setup(phase=air)
    sf_z.set_value_BC(pores=net.pores('top'), values=101325)
    sf_z.set_value_BC(pores=net.pores('bottom'), values=0)
    sf_z.run()
    air.update(sf_z.results())
    K_z = sf_z.calc_effective_permeability(domain_area=area['z'], domain_length=length['z'])[0]
    
    K = [K_x*1e12, K_y*1e12, K_z*1e12]

    return K 

# Connectivity 
conn_unique, conn_counts, conn = find_connectivity(net=net)
avg_connectivity = np.mean(conn)
print(f'Average connectivity: {avg_connectivity :.2f}')

# Pore size distribution
psd = pore_size_distribution(bin_width = 2, pore_volume=net['pore.volume'][net.pores('internal')], 
                             pore_diameter=net['pore.diameter'][net.pores('internal')], pores=net.pores('internal'))
tsd = throat_size_distribution(bin_width = 2, throat_volume=net['throat.volume'][net.throats('internal')], 
                             throat_diameter=net['throat.diameter'][net.throats('internal')], throats=net.throats('internal'))
avg_pore_size = np.mean(net['pore.diameter'][net.pores('internal')])
print(f'Average pore size: {avg_pore_size*1e6 :.2f} um')

# Output to excel
wb = openpyxl.Workbook()
ws = wb.active

output_variables = [conn_unique, conn_counts,  psd['pore_size_range'], 
                    psd['norm_volume'], psd['cum_volume'], tsd['throat_size_range'], 
                    tsd['norm_volume'], tsd['cum_volume']]

output_names = ['pore connections', 'frequency', 'pore size (PSD)', 
                'normalized volume', 'cumulative volume', 'throat size (PSD)', 
                'normalized volume', 'cumulative volume','x length (porosity)',
                'x profile', 'y length', 'yprofile',
                'z length', 'z profile', 'permeability [x,y,z]']

units = ['-', '-', 'um',
         '-', '-', 'um',
         '%', 'um', '%',
         'um', '%', 'Da']

ws.cell(row=1,column=1).value = 'average connectivity (-)'
ws.cell(row=1,column=2).value = avg_connectivity
ws.cell(row=2,column=1).value = 'average pore size (um)'
ws.cell(row=2,column=2).value = avg_pore_size
ws.cell(row=3,column=1).value = 'network porosity (%)'
ws.cell(row=4,column=1).value = 'Internal surface area (cm2/cm3)'
ws.cell(row=5,column=1).value = 'image porosity (%)'

for column_num in range(0, len(output_variables)): # column_num represents the variables in ouput_variables.
    ws.cell(row=1, column=column_num+4).value = output_names[column_num]
    ws.cell(row=2, column=column_num+4).value = units[column_num]
    ws.cell(row=3, column=column_num+4).value = None
    ws.cell(row=4, column=column_num+4).value = None
    for row_num in range(0, len(output_variables[column_num])): # row_num represents the ith entry within the variable array
        ws.cell(row = row_num + 4, column = column_num + 4).value = output_variables[column_num][row_num] # row_num + 5 to convert to Origin format 

wb.save(filename = '.\\output\\' + Type + name + '\\' + 'PSDcounts_' + name + '_network_properties.xlsx')