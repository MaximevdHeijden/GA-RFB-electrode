import pdb                                                  
import openpnm as op
import numpy as np
import time
import openpyxl 
import random
from datetime import datetime
import os
import scipy.optimize as optimize
import math
from numpy import pi as _pi
import GA_Functions_Windows as func
op.Workspace().clear()

''' The main function to run the genetic algorithm:''' 
# Parameters of the Genetic algorithm:
population_size = 50                # Number of initially mutated parents 
offspring_size = population_size    # Number of networks for the offspring generations
num_parents = 10                    # Number of parents used in the mating process
num_generations = 1000              # Number of generations used in the genetic algorithm
fitness_criterion = 1               # Specification of the desired fitness target
mutation_chance = 0.05              # Chance that a mutation occurs in one pore
mutation_range = 0.1                # The fractional offset in pore size that a mutation can induce (e.g. for mutation_range = 0.10, the pore size mutation can cause a mutation in the range 0.90 to 1.10 * the initial pore size)
           
# Initial guess:                       
eta_a = 0.25                        # Anodic overpotential guess
eta_c = -0.25                       # Cathodic overpotential guessSA_Factor = 1                       

# Network properties for cubic networks:
network_shape = [15,15,6]           # The amount of pores in the [x,y,z, direction] - including boundary pores. The width dimension (z-dimension) - defines the thickness of the electrode                                  
spacing = 40e-6                     # Spacing between pores (also defines the maximum size of one pore)
connectivity = 6                    # Connectivity of the cubic network (6: Faces only, 8: Corners only, 12: Edges only, 14: Faces and Corners, 18: Faces and Edges, 20: Edges and Corners, 26: Faces, Edges and Corners)                           
throat_condition = 0.8
ref_pore_diameter = 20e-6
ref_throat_diameter = 15.6e-6

# Merging and splitting conditions:
M_S_chance = 0.999                  # Chance a pore merges/splits
M_S_condition = 0.5                 # Chance the pore(s) merges vs splits. >0.5 the pores are more likely to split.

# Network properties for voronoi networks:
network_shape_2 = [5e-4,5e-4,2.2e-4]
num_points = 300                    # The number of points in the Voronoi structure

# Extracted network:
electrode_name='.\\input\\Freudenberg_FTFF_GA'

# Select your measurement options
# Cubic network (0), extracted network (1), or voronoi network (2):
Extracted = 0
# Merging and splitting off (0) or on (1):
M_and_S = 0
# Mutation off (0) or on (1):
Mut = 1
# Flow-through flow field (0) or interdigitated flow field (1):
FF = 0
if FF == 1:
  network_shape = [28,15,6] 
  network_shape_2 = [10e-4,5e-4,2.2e-4]
  num_points = 600
  electrode_name = '.\\input\\Freudenberg_IDFF_GA'
  # electrode_name=
# Reload a specific individue, no (0) or yes (1):
ReloadData = 0
'''Check in inputdict if the network dimensions are correct, for cubic networks: 012, for extracted networks 201 (use inputDict_GA_V_FH23)'''

# Saving options:
folderold = '.\\GA-2\\Reference'
reload = 900
if ReloadData == 1:
    num_start = reload
else:
    num_start = 0
    reload = 0

name = 'Reference'    
folder = '.\\GA-2\\' + name

# Start of the algorithm:
ws = op.Workspace()

# Import functions and input data: 
if Extracted == 0 or Extracted == 2:
    import inputDict_GA_V as inputDict  
    import customFunctionsGA as cf 
elif Extracted == 1:
    import inputDict_GA_V_FH23 as inputDict  
    import customFunctionsGA_FH23 as cf

# Initialize parent network generation:
net, geo, proj = func.initialize_population(population_size, network_shape, spacing, connectivity, network_shape_2, num_points, ws, throat_condition, electrode_name, mutation_chance, mutation_range, ref_pore_diameter, Extracted, folder)

if ReloadData == 0:
    # Initialize the same kind of electrode as reference network (comparison based on overall volume)
    if Extracted == 0:      
        ref_project=ws.new_project(name='ref_prj')
        ref_network = op.network.Cubic(shape=network_shape, spacing=spacing, connectivity=connectivity, project=ref_project)
        ref_geo = op.geometry.GenericGeometry(network=ref_network, pores=ref_network.Ps, throats=ref_network.Ts, project=ref_project)
        ref_network['pore.internal'][ref_network.pores('surface')] = False     
        op.topotools.trim(network=ref_network, throats=ref_network.throats('surface'))
        health = ref_network.check_network_health()
        op.topotools.trim(network=ref_network, pores=health['trim_pores'])  
        # Add geometry models to the reference network:
        ref_geo['pore.diameter'] = ref_pore_diameter    
        ref_geo['throat.diameter'] = ref_throat_diameter              
        ref_geo = func.add_geometry_models_ref(net=ref_network, geo=ref_geo, throat_condition=throat_condition, spacing=spacing, Extracted=Extracted)
        op.utils.Workspace().save_project(project=ref_project, filename='.\\Genetic_algorithm' + '\\' + folder + '\\j0_1_reference_network')             
    elif Extracted == 2:
        ref_project=ws.new_project(name='ref_prj')
        ref_network=op.network.Voronoi(shape=network_shape_2, num_points=num_points, project=ref_project)
        ref_geo = op.geometry.GenericGeometry(network=ref_network, pores=ref_network.Ps, throats=ref_network.Ts, project=ref_project)
        ref_network['pore.internal'][ref_network.pores('surface')] = False 
        # Add geometry models to the reference network:
        ref_geo['pore.diameter'] = ref_pore_diameter  
        ref_geo = func.add_geometry_models_ref(net=ref_network, geo=ref_geo, throat_condition=throat_condition, spacing=spacing, Extracted=Extracted)
        op.utils.Workspace().save_project(project=ref_project, filename='.\\Genetic_algorithm' + '\\' + folder + '\\j0_1_reference_network')
    else:
        op.utils.Workspace().load_project(filename=electrode_name)
        ref_project = ws[list(ws.keys())[-1]]
        ref_network = ref_project.network
        ref_geo = ref_project['geo_01']
        ref_geo = func.add_geometry_models_ref(net=ref_network, geo=ref_geo, throat_condition=throat_condition, spacing=spacing, Extracted=Extracted)
        op.utils.Workspace().save_project(project=ref_project, filename='.\\Genetic_algorithm' + '\\' + folder + '\\j0_1_reference_network')

    # Check overall volume
    net, geo = func.check_volume(net, geo, ref_network, Extracted, spacing)
    
# Initialize tables:
fitness = np.zeros((num_generations, population_size))
surface = np.zeros((num_generations, population_size))
pumping = np.zeros((num_generations, population_size))
current= np.zeros((num_generations, population_size))

if ReloadData == 1: 
    for ii in range(population_size):
            proj[ii] = op.io.PNM.load_project(filename='.\\Genetic_algorithm' + '\\' + folderold + '\\j0_1_generation_' + str(reload) + '_individual_' + str(ii) + '.pnm')
            net[ii] = proj[ii]['net_01']
            geo[ii] = proj[ii]['geo_01']

for generation in range(num_start, num_generations):
    print("Generation number:", generation)    
    net_copy = {}  
    for i, _ in enumerate(net):
        op.utils.Workspace().save_project(project=net[i].project, filename='.\\Genetic_algorithm' + '\\' + folder + '\\j0_1_generation_' + str(generation) + '_individual_' + str(i))
        op.utils.Workspace().close_project(net[i].project)
    op.Workspace().clear()                      # Prevents the script from cluttering    
    op.utils.Workspace().load_project(filename='.\\Genetic_algorithm' + '\\' + folder + '\\j0_1_reference_network', overwrite=False)
    ref_project = ws[list(ws.keys())[-1]]
    ref_network = ref_project.network
    ref_geo = ref_project['geo_01']
           
    for i, _ in enumerate(net):
        time_start = time.time() 
        # Make a copy of the project and network to have two similar networks. NOTE: Could not find an easier way to copy the network than simply saving and loading i
        op.utils.Workspace().load_project(filename='.\\Genetic_algorithm' + '\\' + folder + '\\j0_1_generation_' + str(generation) + '_individual_' + str(i))
        keys = list(ws.keys())
        proj[i]=ws[keys[2*i+1]]
        net[i] = proj[i].network
        geo[i] = proj[i]['geo_01']    
        op.utils.Workspace().load_project(filename='.\\Genetic_algorithm' + '\\' + folder + '\\j0_1_generation_' + str(generation) + '_individual_' + str(i))
        proj_copy={}
        keys = list(ws.keys())
        proj_copy[i]=ws[keys[-1]]
        net_copy[i] =  proj[i].network     
            
        # Network health diagnosys 
        health = net[i].check_network_health()
        op.topotools.trim(net[i], pores=health['trim_pores'])

        # Run the algorithm for the network and network copy
        P_pump_ideal, P_el, P_el_ideal, P_theory, current_sum, eta_c, eta_a = func.algorithm(net[i], net_copy[i], eta_c, eta_a, generation, Extracted, FF, ReloadData, reload, inputDict, cf, i)        
        # Determine the fitness of the network
        fitness[generation, i] = func.determine_fitness(P_pump_ideal, P_el, P_el_ideal, P_theory)
        surface[generation,i] = net[i]['pore.surface_area'][net[i].pores('internal')].sum()
        current[generation,i] = current_sum
        pumping[generation,i] = P_pump_ideal        
        # print(f'Network number {generation}.{i} converged in {time.time()-time_start} seconds')
    # Display the best result:
    print(f"Highest achieved fitness in generation {generation} is {fitness[generation, :].max()}")
    # Termination criterion
    if fitness.max() > fitness_criterion:
        print(f"Optimal solution found")
        break
   
    # Selection operator
    parent_networks, parent_geometries, parent_projects = func.select_mating_pool(fitness[generation, :], num_parents, net, geo, proj)
    # Save parent geometries
    for i, _ in enumerate(parent_projects):
        op.utils.Workspace().save_project(project=parent_projects[i], filename='.\\Genetic_algorithm' + '\\' + folder + '\\j0_1_generation_' + str(generation) + '_parent_' + str(i))
    # Crossover operators:
    if M_and_S == 0:
        net, geo = func.perform_crossover(parent_networks, offspring_size, generation, Extracted, folder) 
    else:
        net, geo = func.perform_crossover2(parent_networks, offspring_size, spacing, generation, Extracted, folder)             
    # Mutation operator
    if Mut == 1:
        net, geo = func.perform_mutation(net, geo, mutation_chance, mutation_range, Extracted)
    # Add geometry models
    geo = func.add_geometry_models(net, geo, throat_condition, spacing, Extracted)

    # Load reference network:
    op.utils.Workspace().load_project(filename='.\\Genetic_algorithm' + '\\' + folder + '\\j0_1_reference_network', overwrite=False)
    ref_project = ws[list(ws.keys())[-1]]
    ref_network = ref_project.network
    ref_geo = ref_project['geo_01']
      
    # Check overall volume:
    net, geo = func.check_volume(net, geo, ref_network, Extracted, spacing)      
    # Merge and split 
    if M_and_S == 1:
        net, geo = func.split_merge(net, geo, spacing, throat_condition, M_S_chance, M_S_condition, Extracted, ref_network, inputDict)  
    # Network health diagnosys 
    net, geo = func.nw_health_diag(net, geo)
  
    # Save and delete specific generations to save storage capacity: 
    for i, _ in enumerate(net):
        if generation != 0 and generation != 1 and generation != 2 and generation != 3 and generation != 4 and generation != 5 and generation != 6 and generation != 7 and generation != 8 and generation != 9 and generation != 10 and generation%10 != 0 and generation != (num_generations -1):
            os.remove('.\\Genetic_algorithm' + '\\' + folder + "\\j0_1_generation_" + str(generation) + "_individual_" + str(i) + ".pnm")    
    for i, _ in enumerate(parent_projects):
            os.remove('.\\Genetic_algorithm' + '\\' + folder + "\\j0_1_generation_" + str(generation) + "_parent_" + str(i) + ".pnm")      

    if generation == 0 or generation == 1 or generation%100 == 0 or generation == (num_generations -1):  
        # Data storage
        wb = openpyxl.Workbook()
        wba = wb.active
        wba.cell(row=1, column=1).value = 'Generation:'
        wba.cell(row=1, column=2).value = 'Fitness:'
        wba.cell(row=1, column=2+offspring_size).value = 'Maximum Fitness:'
        wba.cell(row=1, column=4+offspring_size).value = 'Surface area:'
        wba.cell(row=1, column=4+offspring_size*2).value = 'Maximum Surface area:'
        wba.cell(row=1, column=2+2*offspring_size+2*2).value = 'Electrical current:'
        wba.cell(row=1, column=2+3*offspring_size+2*2).value = 'Maximum Electrical current:'
        wba.cell(row=1, column=2+3*offspring_size+3*2).value = 'Pumping power:'
        wba.cell(row=1, column=2+4*offspring_size+3*2).value = 'Minimum Pumping power:'
        
        for row_num in range(0, np.shape(fitness)[0]):
            wba.cell(row=row_num+2, column=1).value = row_num    
            for column_num in range(0, np.shape(fitness)[1]):
                wba.cell(row_num+2, column=column_num+2).value = fitness[row_num, column_num]
                wba.cell(row_num+2, column=column_num+4+offspring_size).value = surface[row_num, column_num]
                wba.cell(row_num+2, column=2+column_num+2*offspring_size+2*2).value = current[row_num, column_num]
                wba.cell(row_num+2, column=2+column_num+3*offspring_size+3*2).value = pumping[row_num, column_num]        
                FF_max = fitness.max(1,)
                wba.cell(row=row_num+2, column=2+offspring_size).value = FF_max[row_num]        
                SA_max = surface.max(1,)
                wba.cell(row=row_num+2, column=2+2*offspring_size+2).value = SA_max[row_num]        
                Pel_max = current.max(1,)
                wba.cell(row=row_num+2, column=2+3*offspring_size+2*2).value = Pel_max[row_num]        
                Pp_max = pumping.min(1,)
                wba.cell(row=row_num+2, column=2+4*offspring_size+3*2).value = Pp_max[row_num]        
        now = datetime.now()
        wb.save('Genetic_Algorithm\\' + folder + '\\' + name + '.xlsx')