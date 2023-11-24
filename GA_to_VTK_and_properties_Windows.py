import openpnm as op
import numpy as np
import scipy.optimize as optimize
import customFunctionsGA as cf
import inputDict_GA_V as inputDict_GA
import openpyxl
import time
import os
import matplotlib.pyplot as plt

filepath = '.\\Genetic_Algorithm\\GA-2\\'
electrode_name = (['j0_1_generation_0_individual_0'])
properties = np.zeros((30,5))
properties_2 = np.zeros((30,5,1))
properties_3 = np.zeros((30,5,1))

# Flow-through flow field (0) or interdigitated flow field (1):
FF = 0
# Cubic network (0), extracted network (1), or voronoi network (2):
Extracted = 0

for filename in range(len(electrode_name)):
    # Load in the input dictionary
    param = inputDict_GA.input_dict 
    
    # Retrieve the geometry of the networks
    project_1 = op.io.PNM.load_project(filename=filepath+electrode_name[filename])
    net_c = project_1.network
    project_1 = op.io.PNM.load_project(filename=filepath+electrode_name[filename])
    net_c = project_1.network 
    geom_c = net_c.project['geo_01']  
    project_2 = op.io.PNM.load_project(filename=filepath+electrode_name[filename])# + '_cathode')
    net_a = project_2.network
    geom_a = net_a.project['geo_01']
 
    Type = 'GA\\' 
    os.mkdir('.\\output\\' + Type + electrode_name[filename]) 

    # Instantiate the output Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    polarizationCurveData = {} 
    
    # Define and calculate macro-network properties
    # Electrode properties
    Ps_c = net_c.pores()                                # Pores in cathode
    Ps_a = net_a.pores()                                # Pores in anode
    
     # Define electrode dimensions [x=0, y=1, z=2].
    H_dim = param['height_dimension']                   # The dimension corresponding to the height of the electrodes (sides of the electrode)
    L_dim = param['length_dimension']                   # The dimension corresponding to the length of the electrodes (flow field direction)
    W_dim = param['width_dimension']                    # The dimension corresponding to the width of the electrodes (thickness: current collector -> membrane)
    
    # Calculate the length, height and width of the electrodes
    L_c = np.amax(net_c['pore.coords'][Ps_c, L_dim])  
    H_c = np.amax(net_c['pore.coords'][Ps_c, H_dim])
    H_c_min = np.amin(net_c['pore.coords'][net_c.pores(), H_dim])
    L_c_min = np.amin(net_c['pore.coords'][net_c.pores(), L_dim])
    W_c_min = np.amin(net_c['pore.coords'][net_c.pores(), W_dim])
    W_c = np.amax(net_c['pore.coords'][Ps_c, W_dim])  
    L_a = np.amax(net_a['pore.coords'][Ps_a, L_dim])
    H_a = np.amax(net_a['pore.coords'][Ps_a, H_dim])
    H_a_min = np.amin(net_a['pore.coords'][net_a.pores(), H_dim])
    W_a = np.amax(net_a['pore.coords'][Ps_a, W_dim])    
    L_a_min = np.amin(net_c['pore.coords'][net_c.pores(), L_dim])
    W_a_min = np.amin(net_c['pore.coords'][net_c.pores(), W_dim])

    # Assign the labels 'membrane', 'current collector' 'flow inlet' and 'flow outlet'
    # To the correct boundary pores.
    if FF == 0:
        cf.assign_boundary_poresFF(net_c, W_dim, L_dim, Extracted)
        cf.assign_boundary_poresFF(net_a, W_dim, L_dim, Extracted)    
    elif FF == 1:
        cf.assign_boundary_poresIDFF(net_c, W_dim, L_dim, H_dim, H_c, H_c_min, L_c_min, W_c_min, L_c, W_c, Extracted)
        cf.assign_boundary_poresIDFF(net_a, W_dim, L_dim, H_dim, H_a, H_a_min, L_a_min, W_a_min, L_a, W_a, Extracted) 
    
    A_ext_c = L_c * H_c                                 # Cathode area to compute external current density [m2]
    A_ext_a = L_a * H_a                                 # Anode area [m2]
    mem_area = A_ext_c                                  # Fictitious membrane area [m2]
    if FF == 0:
        A_in_c = W_c * H_c                              # Cathode inlet area [m2]
        A_in_a = W_a * H_a                              # Anode inlet area [m2]
    elif FF == 1:                                       
        # Note that inlet area IDFF is is the thickness of the electrodetimes its length
        A_in_c = W_c * L_c                              # Cathode inlet area [m2]   
        A_in_a = W_a * L_a                              # Anode inlet area [m2]
    
    # Invert anodic network in the width to obtain the same pores at the membrane side for both electrodes:
    # NOTE: The boundaries of the network are also reversed in this direction, see cf.assign_boundary_pores().
    inversion_factor = net_a['pore.coords'][:, W_dim] - min(net_a['pore.coords'][:, W_dim])
    net_a['pore.coords'][:, W_dim]  = max(net_a['pore.coords'][:, W_dim]) - inversion_factor 
     
    # Compute electrochemically active surface area
    Ai_c = net_c['pore.surface_area']                   # Cathode internal surface area [m2]
    Ai_a = net_a['pore.surface_area'] 
    
    # Pore radii
    rp_c = net_c['pore.diameter'] / 2                   # Pore radii in the cathode [m]
    rp_a = net_a['pore.diameter'] / 2                   # Pore radii in the anode [m]
    
    # Create phase and physics objects
    # Anolyte
    anolyte = op.phases.Water(network=net_a, name='anolyte')
    anolyte['pore.electrical_conductivity'] = param['anolyte_conductivity']     
    anolyte['pore.density'] = param['anolyte_density']                          
    anolyte['pore.diffusivity'] = param['D_a']                                  
    anolyte['pore.viscosity'] = param['anolyte_viscosity']     
    phys_a = op.physics.Standard(network=net_a, geometry=geom_a, phase=anolyte)      
    
    # Catholyte
    catholyte = op.phases.Water(network=net_c, name='catholyte')
    catholyte['pore.electrical_conductivity'] = param['catholyte_conductivity'] 
    catholyte['pore.density'] = param['catholyte_density']                      
    catholyte['pore.diffusivity'] = param['D_c']                                
    catholyte['pore.viscosity'] = param['catholyte_viscosity']                  
    phys_c = op.physics.Standard(network=net_c, geometry=geom_c, phase=catholyte)    

    # Regenerate the physics and phase models
    phys_c.regenerate_models()
    phys_a.regenerate_models()
    anolyte.regenerate_models()
    catholyte.regenerate_models()
    
    # Create physics for boundary pores (there should be no resistance to flow or charge transport in the boundary pores):
    if Extracted == 0 or Extracted == 2:    
        boundary_throats_c = net_c.find_neighbor_throats(net_c.pores('surface'))
        boundary_throats_a = net_a.find_neighbor_throats(net_a.pores('surface'))        
        phys_c['throat.ad_dif_conductance'][boundary_throats_c] = 1e4 * phys_c['throat.ad_dif_conductance'].max()
        phys_c['throat.electrical_conductance'][boundary_throats_c] = 1e4 * phys_c['throat.electrical_conductance'].max()        
        phys_a['throat.ad_dif_conductance'][boundary_throats_a] = 1e4 * phys_a['throat.ad_dif_conductance'].max()
        phys_a['throat.electrical_conductance'][boundary_throats_a] = 1e4 * phys_a['throat.electrical_conductance'].max()
        
    # Load in parameters for mass and charge transport    
    # Active species / kinetic parameters
    conc_in_a0 = param['conc_in_a']                     # Anolyte active species inlet concentration in network 0 [mol m-3]
    conc_in_c0 = param['conc_in_c']                     # Catholyte active species inlet concentration in network 0 [mol m-3]
    
    # Cell potential parameters (E_cell < 0 --> Charging, E_cell > 0 --> Discharging)
    E_red_a = param['E_red_a']                          # Standard reduction potential of the anodic half reaction [V]
    E_red_c = param['E_red_c']                          # Standard reduction potential of the cathodic half reaction [V]
    # V_step = param['V_step']                          # Step change in the cell voltage [V] 
    E_cell_final = param['E_cell_final']                # Final value of the cell voltage range [V] 
    
    # Potential parameters
    E_0 = E_red_c - E_red_a                             # Open circuit voltage [V]
    E_cell_vec = [-0.1, -0.2, -0.3, -0.4, -0.5, -0.6, -0.7, -0.8, -0.9, -1]
    
    # Membrane properties
    R_mem = param['membrane_resistivity']               # Flow-through ionic membrane resistivity [Ohm m2] (0.16 Ohm cm2)
    res_mem = R_mem / mem_area                          # Membrane resistance [Ohm]
     
    # Run Stokes Flow algorithm in both half cells for both the odd and the even networks
    r''' Boundary conditions for the even networks are reversed, so that the outlet 
    # pores of e.g. network 1 can be directly mapped to the inlet pores of 
    # network 2 (it are the same pores).'''
    
    sf_c_odd = op.algorithms.StokesFlow(network=net_c, phase=catholyte)
    sf_c_even = op.algorithms.StokesFlow(network=net_c, phase=catholyte)   
    sf_a_odd = op.algorithms.StokesFlow(network=net_a, phase=anolyte)
    sf_a_even = op.algorithms.StokesFlow(network=net_a, phase=anolyte) 
    
    # The inlet fluid velocity is dependent on the local pressure 
    # gradient, which we do not set but follows from the to be computed
    # pressure profile. Hence, we set a target inlet velocity, or 
    # actually a target inlet flowrate, then iteratively update the 
    # inlet pressure to adhere to the target flowrate.
    v_in_c = param['catholyte_inlet_velocity']
    v_in_a = param['anolyte_inlet_velocity']
    Q_in_c = v_in_c * A_in_c                            # total flow rate entering the network [m3/s]
    Q_in_a = v_in_a * A_in_a                            # total flow rate entering the network [m3/s]
    
    if Extracted == 1:
        # Find the inlet throats and compute the inlet area:
        inlet_throats_c_odd = net_c.find_neighbor_throats(pores=net_c.pores('flow_inlet'))
        inlet_throats_a_odd = net_a.find_neighbor_throats(pores=net_a.pores('flow_inlet'))
        outlet_throats_c_odd = net_c.find_neighbor_throats(pores=net_c.pores('flow_outlet'))
        outlet_throats_a_odd = net_a.find_neighbor_throats(pores=net_a.pores('flow_outlet'))
        total_area_inlet_c = net_c['throat.area'][inlet_throats_c_odd].sum()
        total_area_inlet_a = net_a['throat.area'][inlet_throats_a_odd].sum()
        inlet_throats_c_even = net_c.find_neighbor_throats(pores=net_c.pores('flow_outlet'))
        inlet_throats_a_even = net_a.find_neighbor_throats(pores=net_a.pores('flow_outlet'))
        outlet_throats_c_even = net_c.find_neighbor_throats(pores=net_c.pores('flow_inlet'))
        outlet_throats_a_even = net_a.find_neighbor_throats(pores=net_a.pores('flow_inlet'))
        total_area_outlet_c = net_c['throat.area'][outlet_throats_c_even].sum()
        total_area_outlet_a = net_a['throat.area'][outlet_throats_a_even].sum()

        # Artificially increase hydraulic conductance of boundary throats
        Factor_HC_boundary = 1e6
        phys_a['throat.hydraulic_conductance'][inlet_throats_a_odd] = phys_a['throat.hydraulic_conductance'][inlet_throats_a_odd] * Factor_HC_boundary
        phys_c['throat.hydraulic_conductance'][inlet_throats_c_odd] = phys_c['throat.hydraulic_conductance'][inlet_throats_c_odd] * Factor_HC_boundary
        phys_a['throat.hydraulic_conductance'][outlet_throats_a_odd] = phys_a['throat.hydraulic_conductance'][outlet_throats_a_odd] * Factor_HC_boundary
        phys_c['throat.hydraulic_conductance'][outlet_throats_c_odd] = phys_c['throat.hydraulic_conductance'][outlet_throats_c_odd] * Factor_HC_boundary
    else:
        # Find the inlet throats and compute the inlet area:
        inlet_throats_c_odd = net_c.find_neighbor_throats(pores=net_c.pores('flow_inlet'))
        inlet_throats_a_odd = net_a.find_neighbor_throats(pores=net_a.pores('flow_inlet'))   
        inlet_throats_c_even = net_c.find_neighbor_throats(pores=net_c.pores('flow_outlet'))
        inlet_throats_a_even = net_a.find_neighbor_throats(pores=net_a.pores('flow_outlet'))  
        outlet_throats_c_even = net_c.find_neighbor_throats(pores=net_c.pores('flow_inlet'))
        outlet_throats_a_even = net_a.find_neighbor_throats(pores=net_a.pores('flow_inlet'))
        total_area_inlet_c = net_c['throat.area'][inlet_throats_c_odd].sum()
        total_area_inlet_a = net_a['throat.area'][inlet_throats_a_odd].sum()
        total_area_outlet_c = net_c['throat.area'][outlet_throats_c_even].sum()
        total_area_outlet_a = net_a['throat.area'][outlet_throats_a_even].sum()
        
    # Initial guess for the Stokes flow algorithm: 
    # Assign inlet boundary conditions to pores:
    for pore in net_c.pores('flow_inlet'):
        # Boundary pores are connected to one neighboring pore via only one throat
        throat = net_c.find_neighbor_throats(pores=pore)[0]     
        # Neumann inlet boundary condition scaled via the throat connected to the first internal pore
        sf_c_odd.set_rate_BC(values=Q_in_c * net_c['throat.area'][throat] / total_area_inlet_c, pores=pore)    
    for pore in net_a.pores('flow_inlet'):
        throat = net_a.find_neighbor_throats(pores=pore)[0] 
        sf_a_odd.set_rate_BC(values=Q_in_a * net_a['throat.area'][throat] / total_area_inlet_a, pores=pore)    
    for pore in net_c.pores('flow_outlet'):
        throat = net_c.find_neighbor_throats(pores=pore)[0]         
        sf_c_even.set_rate_BC(values=Q_in_c * net_c['throat.area'][throat] / total_area_outlet_c, pores=pore)
    for pore in net_a.pores('flow_outlet'):
        throat = net_a.find_neighbor_throats(pores=pore)[0] 
        sf_a_even.set_rate_BC(values=Q_in_a * net_a['throat.area'][throat] / total_area_outlet_a, pores=pore)       
    
    # Dirichlet outlet boundary condition
    Pout = 0    # Pressure outlet boundary condition [Pa]
    sf_c_odd.set_value_BC(values=Pout, pores=net_c.pores('flow_outlet'))    # Dirichlet outlet boundary condition
    sf_c_even.set_value_BC(values=Pout, pores=net_c.pores('flow_inlet'))    # Dirichlet outlet boundary condition
    sf_a_odd.set_value_BC(values=Pout, pores=net_a.pores('flow_outlet'))    # Dirichlet outlet boundary condition
    sf_a_even.set_value_BC(values=Pout, pores=net_a.pores('flow_inlet'))    # Dirichlet outlet boundary condition
    
    # Run electrolyte transport algorithms
    sf_c_odd.run()
    sf_c_even.run()
    sf_a_odd.run()
    sf_a_even.run()
    
    # Update phases with calculated pressure field
    # The phases are updated within the iterative algorithm depending on the network number
    catholyte.update(sf_c_even.results())
    anolyte.update(sf_a_even.results())
    
    # Find the inlet pressure, so that the target flow rate is reached
    def inlet_pressure_c(P_in, sf_c, Q_desired, inlet_throats, inlet_pores, net, phase, phys):
        
        sf_c.set_value_BC(values=P_in, pores=inlet_pores) # Dirichlet boundary inlet condition
        sf_c.run()
        
        phase.update(sf_c.results())
        
        pore_1 = net['throat.conns'][inlet_throats][:, 0]
        pore_2 = net['throat.conns'][inlet_throats][:, 1]
        
        delta_P = abs(phase['pore.pressure'][pore_1] - phase['pore.pressure'][pore_2])
        Q_tot = (phys['throat.hydraulic_conductance'][inlet_throats] * delta_P).sum()
    
        return (Q_tot - Q_desired)
    
    def inlet_pressure_a(P_in, sf_a, Q_desired, inlet_throats, inlet_pores, net, phase, phys):
        
        sf_a.set_value_BC(values=P_in, pores=inlet_pores) # Dirichlet boundary inlet condition
        sf_a.run()
        
        phase.update(sf_a.results())
        
        pore_1 = net['throat.conns'][inlet_throats][:, 0]
        pore_2 = net['throat.conns'][inlet_throats][:, 1]
        
        delta_P = abs(phase['pore.pressure'][pore_1] - phase['pore.pressure'][pore_2])
        Q_tot = (phys['throat.hydraulic_conductance'][inlet_throats] * delta_P).sum()
    
        return (Q_tot - Q_desired)

    # Set initial guesses to the mean pressure found above
    x0_c_even = catholyte['pore.pressure'][net_c.pores('flow_outlet')].mean()
    x0_a_even = anolyte['pore.pressure'][net_a.pores('flow_outlet')].mean()
    
    # The phases are updated within the iterative algorithm depending on the network number
    catholyte.update(sf_c_odd.results())
    anolyte.update(sf_a_odd.results())
    
    # Set initial guesses to the mean pressure found above
    x0_c_odd = catholyte['pore.pressure'][net_c.pores('flow_inlet')].mean()
    x0_a_odd = anolyte['pore.pressure'][net_a.pores('flow_inlet')].mean()
    
    # Delete the initial guess boundary condition
    sf_c_odd['pore.bc_rate'] = np.nan
    sf_c_even['pore.bc_rate'] = np.nan
    sf_a_odd['pore.bc_rate'] = np.nan    
    sf_a_even['pore.bc_rate'] = np.nan    
    
    # Find the pressure at the inlet at which the total flow rate matches the desired flow rate.
    optimize.fsolve(func=inlet_pressure_c, x0=x0_c_odd, args=(sf_c_odd, Q_in_c, inlet_throats_c_odd, net_c.pores('flow_inlet'), net_c, catholyte, phys_c))
    optimize.fsolve(func=inlet_pressure_a, x0=x0_a_odd, args=(sf_a_odd, Q_in_a, inlet_throats_a_odd, net_a.pores('flow_inlet'), net_a, anolyte, phys_a))
    optimize.fsolve(func=inlet_pressure_c, x0=x0_c_even, args=(sf_c_even, Q_in_c, inlet_throats_c_even, net_c.pores('flow_outlet'), net_c, catholyte, phys_c))
    optimize.fsolve(func=inlet_pressure_a, x0=x0_a_even, args=(sf_a_even, Q_in_a, inlet_throats_a_even, net_a.pores('flow_outlet'), net_a, anolyte, phys_a))
    
    # Set up iterative algorithm for mass and charge transport
    # Set up advection diffusion algorithms
    ad_c_odd = op.algorithms.AdvectionDiffusion(network=net_c, phase=catholyte)
    ad_c_even = op.algorithms.AdvectionDiffusion(network=net_c, phase=catholyte)
    ad_a_odd = op.algorithms.AdvectionDiffusion(network=net_a, phase=anolyte)
    ad_a_even = op.algorithms.AdvectionDiffusion(network=net_a, phase=anolyte)
    
    # Boundary conditions
    # The inlet boundary conditions are updated in the iterative loop (conc_in of network 2 = conc_out of network 1).
    ad_c_odd.set_outflow_BC(pores=net_c.pores('flow_outlet'))
    ad_c_even.set_outflow_BC(pores=net_c.pores('flow_inlet'))
    ad_a_odd.set_outflow_BC(pores=net_a.pores('flow_outlet'))
    ad_a_even.set_outflow_BC(pores=net_a.pores('flow_inlet'))
    
    # Source term
    # It should always be a 'string' loaded into the model parameters, no 0.
    # regen_mode = "deferred" bypasses OpenPNM warning that concentration is yet undefined
    source_term = op.models.physics.generic_source_term.linear
    
    phys_a['pore.rxn_A2'] = 0.0
    phys_c['pore.rxn_A2'] = 0.0
    
    anolyte.add_model(propname='pore.butler_volmer', 
                      model=source_term,
                      A1='pore.rxn_A1', A2="pore.rxn_A2",
                      X='pore.concentration',
                      regen_mode="deferred")
    
    catholyte.add_model(propname='pore.butler_volmer', 
                        model=source_term,
                        A1='pore.rxn_A1', 
                        A2="pore.rxn_A2", 
                        X='pore.concentration',
                        regen_mode="deferred")
    
    ad_c_odd.set_source(propname='pore.butler_volmer', pores=net_c.pores('internal'))
    ad_c_even.set_source(propname='pore.butler_volmer', pores=net_c.pores('internal'))
    ad_a_odd.set_source(propname='pore.butler_volmer', pores=net_a.pores('internal'))
    ad_a_even.set_source(propname='pore.butler_volmer', pores=net_a.pores('internal'))
    
    # Set up the ohmic conduction algorithm
    oc_c = op.algorithms.OhmicConduction(network=net_c, phase=catholyte)
    oc_a = op.algorithms.OhmicConduction(network=net_a, phase=anolyte)
    
    # Boundary conditions
    r"""
    Boundary condition at membrane is a function of the current and can therefore
    be found in the iteration loop, other boundaries are automatically set to
    no flux, i.e. dV/dx = 0
    """
    
    # Source term
    # Initial value for A1 is 1 and should be zeros, function can be linearized later on to improve stability.
    # regen_mode = "deferred" bypasses OpenPNM warning that voltage is yet undefined
    phys_a['pore.current_A1'] = 0.0
    phys_c['pore.current_A1'] = 0.0
    
    anolyte.add_model(propname='pore.proton_anolyte', 
                      model=source_term,
                      A1='pore.current_anolyte_A1',
                      A2='pore.current_anolyte_A2',
                      X='pore.voltage',
                      regen_mode="deferred")
    
    catholyte.add_model(propname='pore.proton_catholyte', 
                        model=source_term,
                        A1='pore.current_catholyte_A1',
                        A2='pore.current_catholyte_A2',
                        X='pore.voltage',
                        regen_mode="deferred")
    
    oc_c.set_source(propname='pore.proton_catholyte', pores=net_c.pores('internal'))
    oc_a.set_source(propname='pore.proton_anolyte', pores=net_a.pores('internal'))
    
    # Compute the coupled charge and mass transfer equations
    number_of_networks = 1#int(np.ceil(param['total_electrode_length']/L_c)) # Calculate the total required networks in series for the desired network length.
    
    # Initialize overpotential and concentration guesses
    eta0_a = np.zeros((net_a.Np, number_of_networks))
    eta0_c = np.zeros((net_c.Np, number_of_networks))
    
    # Initialize concentration in pores to inlet concentration
    conc0_a = np.ones((net_a.Np, number_of_networks)) * conc_in_a0 
    conc0_c = np.ones((net_c.Np, number_of_networks)) * conc_in_c0
    
    # Matrixes to store inlet concentration for the odd and even networks
    conc0_in_a_odd = np.zeros((len(net_a.pores('flow_inlet')), number_of_networks))
    conc0_in_c_odd = np.zeros((len(net_c.pores('flow_inlet')), number_of_networks))
    conc0_in_a_even = np.zeros((len(net_a.pores('flow_outlet')), number_of_networks))
    conc0_in_c_even = np.zeros((len(net_c.pores('flow_outlet')), number_of_networks))
    
    # Set the inlet concentration of the 0th network to the inlet concentration of the electrode
    conc0_in_a_odd[:, 0] = conc_in_a0
    conc0_in_c_odd[:, 0] = conc_in_c0
    
    rel_tol = param['rel_tol']                          # Relative tolerance
    abs_tol = param['abs_tol']                          # Absolute tolerance for external current density
    max_iter = param['max_iter']                        # Maximum number of iterations
    omega = param['omega']                              # Damping factor

    for E_cell_idx, E_cell in enumerate(E_cell_vec):
        for network in range(number_of_networks):
             # Update guesses for the overpotential and concentration based on the previous result
            eta_c = eta0_c[:, network]                  # Overpotential [V]
            eta_a = eta0_a[:, network]
    
            if (network + 1) % 2 == 0: # Check if the network is odd or even to set inlet concentration
                conc_in_a = conc0_in_a_even[:, network] 
                conc_in_c = conc0_in_c_even[:, network]
            else:
                conc_in_a = conc0_in_a_odd[:, network]
                conc_in_c = conc0_in_c_odd[:, network]
    
            conc_c = conc0_c[:, network]                # Set concentration in all pores based on previous result [mol m-3]
            conc_a = conc0_a[:, network]
            
            # The current is reset, to reset the relative error
            current_ad_a = 0.0                          # Current [A]
            current_oc_a = 0.0
            current_ad_c = 0.0
            current_oc_c = 0.0
            
            V_c = E_cell - E_0 - eta_c                  # Voltage in the liquid phase [V]
            V_a = 0 - eta_a
            
            anolyte['pore.voltage'] = V_a
            catholyte['pore.voltage'] = V_c
            
            # Flip the network for visualization in paraview
            inversion_factor = net_a['pore.coords'][:, L_dim] - min(net_a['pore.coords'][:, L_dim])
            new_position = max(net_a['pore.coords'][:, L_dim]) - inversion_factor
            net_a['pore.coords'][:, L_dim] = new_position
            net_c['pore.coords'][:, L_dim] = new_position
    
            # Iterative loop to calculate coupled charge and mass transfer equations
            for itr in range(max_iter):
                # Cathodic half cell
                # ------------------------------------------------------------------------------------------------------------------
                # Update ad reaction term with new values for eta_c
                phys_c['pore.rxn_A1'] = cf.bv_rate_constant_ad_c(eta_c, Ai_c, rp_c)
    
                # network+1 for the network number since the range function starts counting at network 0
                if (network + 1) % 2 == 0:
                    # Retrieve the StokesFlow data for an even network
                    catholyte.update(sf_c_even.results())
                    
                    # Regenerate (the ad_dif_conductance) models
                    phys_c.regenerate_models('throat.ad_dif_conductance')
                    catholyte.regenerate_models('throat.peclet.ad')
                    
                    # Set the boundary conditions for even networks (reverted case)
                    ad_c_even.set_value_BC(values=conc_in_c, pores=net_c.pores('flow_outlet'))
                    ad_c_even.set_outflow_BC(pores=net_c.pores('flow_inlet'))
                    
                    # Solve the advection-diffusion-reaction equation
                    ad_c_even.run()
                    
                    # Update the concentration
                    conc_new_c = ad_c_even['pore.concentration']
                else:
                    # Retrieve the StokesFlow data for an even network
                    catholyte.update(sf_c_odd.results())
                    
                    # Regenerate the ad_dif_conductance model
                    phys_c.regenerate_models('throat.ad_dif_conductance')
                    catholyte.regenerate_models('throat.peclet.ad')
                    
                    # Set the boundary conditions for odd networks (normal case)
                    ad_c_odd.set_value_BC(values=conc_in_c, pores=net_c.pores('flow_inlet'))
                    ad_c_odd.set_outflow_BC(pores=net_c.pores('flow_outlet'))
                    
                    # Solve the advection-diffusion-reaction equation
                    ad_c_odd.run()
                    
                    # Update the concentration
                    conc_new_c = ad_c_odd['pore.concentration']
               
                # Update the value for the concentration
                conc_c = conc_new_c * omega + conc_c * (1 - omega)
                catholyte['pore.concentration'] = conc_c
                
                # Compute current according to the reaction of species
                # NOTE: we only use the internal pores for the relative error calculation, 
                # NOTE: as both the ohmic conduction and advection-diffusion algorithm use different boundary pores.
                current_estimation_ad_c = cf.bv_rate_constant_oc_c(conc_c, eta_c, Ai_c, rp_c)[net_c.pores('internal')].sum()
                
                # Compute oc reaction term with new values for conc_c
                drdv = cf.bv_rate_derivative_oc_c(conc_c, eta_c, Ai_c, rp_c)
                r = cf.bv_rate_constant_oc_c(conc_c, eta_c, Ai_c, rp_c)
                
                phys_c['pore.current_catholyte_A1'] = drdv
                phys_c['pore.current_catholyte_A2'] = r - drdv * V_c
    
                # Couple the potential field in both half cells via a continuum averaged membrane approach
                # This only works when identical anodic and cathodic networks are used (networks with identical membrane pores)
                # (the membrane pores remain on the membrane side for both odd and even networks)
                V_membrane_pores_a = anolyte['pore.voltage'][net_a.pores('membrane')]
                V_membrane_pores_c = V_membrane_pores_a + res_mem * current_estimation_ad_c  # [V]
                
                # Update the membrane boundary condition for the cathode
                oc_c.set_value_BC(values=V_membrane_pores_c, pores=net_c.pores('membrane'))
                
               # Solve the ohmic conduction equation (potential field) for the liquid electrolyte in the porous cathode
                oc_c.run()
    
                # Update the value for the liquid potential
                V_new_c = oc_c['pore.voltage']
                V_c = V_new_c * omega + V_c * (1 - omega)
                catholyte['pore.voltage'] = V_c
                
                # Compute current according to the transport of species
                current_estimation_oc_c = cf.bv_rate_constant_oc_c(conc_c, eta_c, Ai_c, rp_c)[net_c.pores('internal')].sum()
    
                # Compute new cathodic overpotential
                eta_c = E_cell - V_c - E_0
                
                # Anodic half cell
                # Update reaction term with new values for eta
                phys_a['pore.rxn_A1'] = cf.bv_rate_constant_ad_a(eta_a, Ai_a, rp_a)
    
                # Solve the advection-diffusion-reaction equation for the anode
                if (network + 1) % 2 == 0:
                    # Retrieve the StokesFlow data for an even network
                    anolyte.update(sf_a_even.results())
                    
                    # Regenerate the (ad_dif_conductance) models
                    phys_a.regenerate_models('throat.ad_dif_conductance')
                    anolyte.regenerate_models('throat.peclet.ad')
                    
                    # Set the boundary conditions for even networks (reverted case)
                    ad_a_even.set_value_BC(values=conc_in_a, pores=net_a.pores('flow_outlet'))
                    ad_a_even.set_outflow_BC(pores=net_a.pores('flow_inlet'))
                    
                    # Solve the advection-diffusion-reaction equation
                    ad_a_even.run()
                    
                    # Update the concentration
                    conc_new_a = ad_a_even['pore.concentration']
                else:
                    # Retrieve the StokesFlow data for an odd network
                    anolyte.update(sf_a_odd.results())
                    
                    # Regenerate the (ad_dif_conductance) models
                    phys_a.regenerate_models('throat.ad_dif_conductance')
                    anolyte.regenerate_models('throat.peclet.ad')
                    
                    # Set the boundary conditions for odd networks (normal case)
                    ad_a_odd.set_value_BC(values=conc_in_a, pores=net_a.pores('flow_inlet'))
                    ad_a_odd.set_outflow_BC(pores=net_a.pores('flow_outlet'))
                    
                    # Solve the advection-diffusion-reaction equation
                    ad_a_odd.run()
                    
                    # Update the concentration
                    conc_new_a = ad_a_odd['pore.concentration']
    
                # Update the value for the concentration
                conc_a = conc_new_a * omega + conc_a * (1 - omega)
                anolyte['pore.concentration'] = conc_a
                
                # Compute current according to the reaction of species
                current_estimation_ad_a = cf.bv_rate_constant_oc_a(conc_a, eta_a, Ai_a, rp_a)[net_a.pores('internal')].sum()
    
                # Compute oc reaction term with new values for conc_c
                # The current generation is a source instead of a sink term, so minus bv_rate_constant_oc_anode
                drdv = cf.bv_rate_derivative_oc_c(conc_a, eta_a, Ai_a, rp_a)
                r = -cf.bv_rate_constant_oc_a(conc_a, eta_a, Ai_a, rp_a)
                
                phys_a['pore.current_anolyte_A1'] = drdv
                phys_a['pore.current_anolyte_A2'] = r - drdv * V_a
    
                # Couple the potential field in both half cells via the membrane pores
                V_membrane_pores_c = catholyte['pore.voltage'][net_c.pores('membrane')]
                V_membrane_pores_a = V_membrane_pores_c - res_mem * current_estimation_ad_a  # [V]
                
                # Update the membrane boundary condition for the anode
                oc_a.set_value_BC(values=V_membrane_pores_a, pores=net_a.pores('membrane'))
                
                # Solve the ohmic conduction equation (potential field) for the liquid electrolyte in the porous cathode
                oc_a.run()
    
                # Update voltage
                V_new_a = oc_a['pore.voltage']
                V_a = V_new_a * omega + V_a * (1 - omega)
                anolyte['pore.voltage'] = V_a
    
                # Compute new anodic overpotentials
                eta_a = 0 - V_a
                
                # Compute current according to the transport of species
                current_estimation_oc_a = cf.bv_rate_constant_oc_a(conc_a, eta_a, Ai_a, rp_a)[net_a.pores('internal')].sum()
    
                # Convergence and updating
                # Calculate the relative error with the previous solution
                rel_err_current_ad_a = cf.rel_error(current_estimation_ad_a, current_ad_a)
                rel_err_current_oc_a = cf.rel_error(current_estimation_oc_a, current_oc_a)
                rel_err_current_ad_c = cf.rel_error(current_estimation_ad_c, current_ad_c)
                rel_err_current_oc_c = cf.rel_error(current_estimation_oc_c, current_oc_c)
                
                # Store the previous solution
                current_ad_a = current_estimation_ad_a
                current_oc_a = current_estimation_oc_a
                current_ad_c = current_estimation_ad_c
                current_oc_c = current_estimation_oc_c
    
                # Calculate the absolute error between the current density found in the
                # anodic and cathodic compartment [A cm-2] (1e4 = conversion m2 to cm2)
                abs_err_current_cat_an = abs((current_oc_a / A_ext_a / 1e4 - current_oc_c / A_ext_c / 1e4))
                
                # Check if found progression of the solution is within tolerances
                convergence_ad_a = rel_err_current_ad_a < rel_tol
                convergence_oc_a = rel_err_current_oc_a < rel_tol
                convergence_ad_c = rel_err_current_ad_c < rel_tol
                convergence_oc_c = rel_err_current_oc_c < rel_tol
                convergence_a_c = abs_err_current_cat_an < abs_tol
                
                if (convergence_ad_a and convergence_oc_a and convergence_ad_c and convergence_oc_c and convergence_a_c or (abs(current_estimation_ad_c)+abs(current_estimation_ad_a))/1e4/A_ext_c < 1e-5 and itr > 20) or itr == max_iter-1:
                    print(f"CONVERGED network {network+1} at {E_cell} [V]  with {current_estimation_ad_c/A_ext_c/1e4:.5f} [A/cm2]!")
    
                    # Compute total amount of current from all species that is consumed (in all pores except for the the flow boundary pores)
                    # Post-treatment
                    # Current in every pore
                    anolyte['pore.current'] = cf.bv_rate_constant_oc_a(conc_a, eta_a, Ai_a, rp_a)
                    catholyte['pore.current'] = cf.bv_rate_constant_oc_c(conc_c, eta_c, Ai_c, rp_c)
    
                    # Overpotential contributions
                    # NOTE: In a symmetric cell setup, the applied voltage is equal to the total overpotential within the cell.               
                    # Ohmic overpotential contribution:
                    catholyte['pore.ohmic_overpotential'] = V_c - V_membrane_pores_c.mean()
                    anolyte['pore.ohmic_overpotential'] = V_a  - V_membrane_pores_a.mean()
                    eta_ohm_mem = V_membrane_pores_c.mean()-V_membrane_pores_a.mean()
    
                    # Activation overpotential contribution:
                    eta_act_c = np.zeros(net_c.Np)
                    eta_act_a = np.zeros(net_a.Np)
                
                    for pore in range(net_c.Np):
                        eta_act_c[pore] = optimize.fsolve(func = cf.find_eta_act, x0=eta_c[pore], args=('cathode', catholyte['pore.current'][pore], Ai_c[pore]))
                        # eta_act_c[pore] = optimize.least_squares(fun=cf.find_eta_act, bounds=(eta_c[pore],0), x0=eta_c[pore], args=('cathode', catholyte['pore.current'][pore], Ai_c[pore])).x
        
                    for pore in range(net_a.Np):
                        eta_act_a[pore] = optimize.fsolve(func = cf.find_eta_act, x0=eta_a[pore], args=('anode', anolyte['pore.current'][pore], Ai_a[pore]))
                        #eta_act_a[pore] = optimize.least_squares(fun=cf.find_eta_act, bounds=(0, eta_a[pore]), x0=eta_a[pore], args=('anode', anolyte['pore.current'][pore], Ai_a[pore])).x
                    
                    catholyte['pore.activation_overpotential'] = eta_act_c
                    anolyte['pore.activation_overpotential'] = eta_act_a
                    
                    # Concentration overpotential contribution
                    catholyte['pore.concentration_overpotential'] = eta_c-catholyte['pore.activation_overpotential'] 
                    anolyte['pore.concentration_overpotential'] = eta_a-anolyte['pore.activation_overpotential']
    
                    # Conpute total current
                    # NOTE: Boundary pores are excluded, since these are artificial pores
                    # added by the SNOW algorithm, that are not part of the real network
                    current_sum_a = anolyte['pore.current'][net_a.pores('internal')].sum()
                    current_sum_c = catholyte['pore.current'][net_c.pores('internal')].sum()
                    current_sum = (abs(current_sum_a)+abs(current_sum_c))/2
    
                    # Output to Excel workbook
                    polarizationCurveData['current density'] = current_sum/A_ext_a/1e4
                    polarizationCurveData['cell voltage'] = abs(E_cell)
                    polarizationCurveData['anodic activation overpotential'] = anolyte['pore.activation_overpotential'][net_a.pores('internal')].mean() 
                    polarizationCurveData['cathodic activation overpotential'] = catholyte['pore.activation_overpotential'][net_c.pores('internal')].mean()
                    polarizationCurveData['anodic concentration overpotential'] = anolyte['pore.concentration_overpotential'][net_a.pores('internal')].mean()
                    polarizationCurveData['cathodic concentration overpotential'] = catholyte['pore.concentration_overpotential'][net_c.pores('internal')].mean()
                    polarizationCurveData['anodic ohmic overpotential'] = anolyte['pore.ohmic_overpotential'][net_a.pores('internal')].mean()
                    polarizationCurveData['cathodic ohmic overpotential'] = catholyte['pore.ohmic_overpotential'][net_c.pores('internal')].mean()
                    polarizationCurveData['membrane ohmic overpotential'] = abs(eta_ohm_mem)           
                    
                    for idx, key in enumerate(polarizationCurveData):
                        ws.cell(row=1, column=idx+1).value = key
                        ws.cell(row=4+network+number_of_networks*E_cell_idx, column=idx+1).value = polarizationCurveData[key]
                                        
                    # Update initial guesses for next cell voltage
                    eta0_a[:, network] = eta_a
                    eta0_c[:, network] = eta_c
                    conc0_a[:, network] = conc_a
                    conc0_c[:, network] = conc_c
    
                    if network is not number_of_networks-1:
                        if (network+1) % 2 == 0:
                            conc0_in_a_odd[:, network+1] = conc_a[net_a.pores('flow_inlet')]
                            conc0_in_c_odd[:, network+1] = conc_c[net_c.pores('flow_inlet')]
                        else:
                            conc0_in_a_even[:, network+1] = conc_a[net_a.pores('flow_outlet')]
                            conc0_in_c_even[:, network+1] = conc_c[net_c.pores('flow_outlet')]
    
                    # The paraview output can be optimized by incrementing the network coordinates with the
                    # network length for every network and by inverting the network over the length direction
                    if (E_cell_idx+1)%10 == 0:
                        op.io.VTK.save(network=net_c, phases=catholyte, filename='.\\output\\' + Type + '\\' + electrode_name[filename] + '\\cathode_' + str(E_cell)[0:2] + '_' + str(E_cell)[3:] + 'V_' + 'network' + str(network))  # Varying on your OS this output can cause an error, due to the decimal in E_cell
                        op.io.VTK.save(network=net_a, phases=anolyte, filename='.\\output\\' + Type + '\\' + electrode_name[filename] + '\\anode_' + str(E_cell)[0:2] + '_' + str(E_cell)[3:] + 'V'+ 'network' + str(network))
                    break
       
    wb.save(filename = '.\\output\\' + Type + electrode_name[filename] + '\\' + electrode_name[filename] +'polarizationCurveData.xlsx')
               
    net_c['throat.radius'] = net_c['throat.diameter']/2
    net_a['throat.radius'] = net_a['throat.diameter']/2
    
    op.io.VTK.save(network=net_c, filename=filepath+electrode_name[filename]+'_3D_cathode', phases=net_c.project['catholyte'])
    op.io.VTK.save(network=net_a, filename=filepath+electrode_name[filename]+'_3D_anode', phases=net_a.project['anolyte'])
    
    
    # =============================================================================
    # Plot length-thickness hexagonal binning plots for the current field
    # Make 2D network for report:
    net_c['pore.layer_Internal_mem'] = False 
    net_c['pore.layer_Internal_cc'] = False 
    net_c['pore.layer_mem'] = False 
    net_c['pore.layer_cc'] = False
    net_c['throat.diameter_2'] = net_c['throat.diameter'] * 1e6
    
    net_c['throat.layer_Internal_mem'] = False 
    net_c['throat.layer_Internal_cc'] = False 
    net_c['throat.layer_mem'] = False 
    net_c['throat.layer_cc'] = False
    
    net_c['throat.im_down'] = False 
    net_c['throat.icc_down'] = False 
    net_c['throat.m_down'] = False 
    net_c['throat.cc_down'] = False
       
    for pore in range(net_c.Np):
        if 6e-5*0.99 < net_c['pore.coords'][pore, 2] < 6e-5*1.01:
            net_c['pore.layer_cc'][pore] = True
    
    for pore in range(net_c.Np):
        if 1e-4*0.9 < net_c['pore.coords'][pore, 2] < 1e-4*1.1:
            net_c['pore.layer_Internal_cc'][pore] = True    
        
    for pore in range(net_c.Np):
        if 1.4e-4*0.99 < net_c['pore.coords'][pore, 2] < 1.4e-4*1.01:
            net_c['pore.layer_Internal_mem'][pore] = True
    
    for pore in range(net_c.Np):
        if 1.8e-4*0.99 < net_c['pore.coords'][pore, 2] < 1.8e-4*1.01:
            net_c['pore.layer_mem'][pore] = True
            
    pores = np.setdiff1d(net_c.pores('all'), net_c.pores('layer_mem'))
    #op.topotools.trim(network=net_c, pores=pores)  
    op.io.VTK.save(network=net_c, filename=filepath+electrode_name[filename]+'_3D_cathode_parts', phases=net_c.project['catholyte'])
    op.io.VTK.save(network=net_a, filename=filepath+electrode_name[filename]+'_3D_anode_parts', phases=net_a.project['anolyte'])
    op.io.VTK.save(network=net_c, filename=filepath+electrode_name[filename]+'_2D_mem', phases=net_c.project['catholyte'])
     
    # =============================================================================
    # CREATING NEW TAGS FOR MEM, CC, ICC AND IM LAYERS   
    # INTERNAL LAYER NEXT TO CURRENT COLLECTOR
    im = net_c.get_incidence_matrix(fmt='coo')
    neighbor_pores = net_c.find_neighbor_pores(net_c.pores('layer_Internal_cc'), mode='union')
    all_throats = net_c.find_neighbor_throats(net_c.pores('layer_Internal_cc'), mode='union')
    ex_throats = net_c.find_neighbor_throats(neighbor_pores, mode='union')
    internal_cc_throats = np.setdiff1d(all_throats,ex_throats)
    net_c['throat.layer_Internal_cc'][internal_cc_throats] = True
    
    conn_icc = net_c.find_connected_pores(throats=internal_cc_throats) 
    for i in range(len(internal_cc_throats)):
        if not -0.00001 < (net_c['pore.coords'][conn_icc[i,0]] - net_c['pore.coords'][conn_icc[i,1]])[1] < 0.00001:
            net_c['throat.icc_down'][internal_cc_throats[i]] = True
            
    # INTERNAL LAYER NEXT TO MEMBRANE
    neighbor_pores = net_c.find_neighbor_pores(net_c.pores('layer_Internal_mem'), mode='union')
    all_throats = net_c.find_neighbor_throats(net_c.pores('layer_Internal_mem'), mode='union')
    ex_throats = net_c.find_neighbor_throats(neighbor_pores, mode='union')
    internal_mem_throats = np.setdiff1d(all_throats,ex_throats)
    net_c['throat.layer_Internal_mem'][internal_mem_throats] = True
    
    conn_im = net_c.find_connected_pores(throats=internal_mem_throats) 
    for i in range(len(internal_mem_throats)):
        if not -0.00001 < (net_c['pore.coords'][conn_im[i,0]] - net_c['pore.coords'][conn_im[i,1]])[1] < 0.00001:
            net_c['throat.im_down'][internal_mem_throats[i]] = True
    
    # LAYER NEXT TO CURRENT COLLECTOR
    neighbor_pores = net_c.find_neighbor_pores(net_c.pores('layer_cc'), mode='union')
    all_throats = net_c.find_neighbor_throats(net_c.pores('layer_cc'), mode='union')
    ex_throats = net_c.find_neighbor_throats(neighbor_pores, mode='union')
    cc_throats = np.setdiff1d(all_throats,ex_throats)
    net_c['throat.layer_cc'][cc_throats] = True
    
    conn_cc = net_c.find_connected_pores(throats=cc_throats) 
    for i in range(len(cc_throats)):
        if not -0.00001 < (net_c['pore.coords'][conn_cc[i,0]] - net_c['pore.coords'][conn_cc[i,1]])[1] < 0.00001:
            net_c['throat.cc_down'][cc_throats[i]] = True
    
    # LAYER NEXT TO MEMBRANE
    neighbor_pores = net_c.find_neighbor_pores(net_c.pores('layer_mem'), mode='union')
    all_throats = net_c.find_neighbor_throats(net_c.pores('layer_mem'), mode='union')
    ex_throats = net_c.find_neighbor_throats(neighbor_pores, mode='union')
    mem_throats = np.setdiff1d(all_throats,ex_throats)
    net_c['throat.layer_mem'][mem_throats] = True
    
    conn_mem = net_c.find_connected_pores(throats=mem_throats) 
    for i in range(len(mem_throats)):
        if not -0.00001 < (net_c['pore.coords'][conn_mem[i,0]] - net_c['pore.coords'][conn_mem[i,1]])[1] < 0.00001:
            net_c['throat.m_down'][mem_throats[i]] = True
            
    # =============================================================================
    # SAVING DATA    
    op.io.VTK.save(network=net_c, phases=catholyte, filename='.\\output\\' + Type + electrode_name[filename] + '\\' +'cathode_' + 'surface')
    properties[0,filename] = net_c['pore.diameter'].mean()
    properties[1,filename] = net_c['throat.diameter'].mean()  
    properties[2,filename] = net_c['pore.surface_area'][net_c.pores('layer_cc')].sum() #mean()
    properties[3,filename] = net_c['pore.surface_area'][net_c.pores('layer_Internal_cc')].sum()#mean()
    properties[4,filename] = net_c['pore.surface_area'][net_c.pores('layer_Internal_mem')].sum()#mean()
    properties[5,filename] = net_c['pore.surface_area'][net_c.pores('layer_mem')].sum() #mean()
    properties[6,filename] = net_c['throat.diameter'][net_c.throats('layer_cc')].mean()
    properties[7,filename] = net_c['throat.diameter'][net_c.throats('layer_Internal_cc')].mean()
    properties[8,filename] = net_c['throat.diameter'][net_c.throats('layer_Internal_mem')].mean()
    properties[9,filename] = net_c['throat.diameter'][net_c.throats('layer_mem')].mean()
    properties[10,filename] = net_c['pore.diameter'][net_c.pores('layer_cc')].mean()
    properties[11,filename] = net_c['pore.diameter'][net_c.pores('layer_Internal_cc')].mean()
    properties[12,filename] = net_c['pore.diameter'][net_c.pores('layer_Internal_mem')].mean()
    properties[13,filename] = net_c['pore.diameter'][net_c.pores('layer_mem')].mean()
    properties_2[1,filename,:] = net_c['pore.surface_area'][net_c.pores('layer_cc')].sum()
    properties_2[2,filename,:] = net_c['pore.surface_area'][net_c.pores('layer_Internal_cc')].sum()
    properties_2[3,filename,:] = net_c['pore.surface_area'][net_c.pores('layer_Internal_mem')].sum()
    properties_2[4,filename,:] = net_c['pore.surface_area'][net_c.pores('layer_mem')].sum()
    
    inlet_throats_c = net_c.find_neighbor_throats(pores=net_c.pores('surface'))
    op.topotools.trim(network=net_c, throats=inlet_throats_c)
    
    properties_3[1,filename,:] = net_c['throat.diameter'][net_c.throats('layer_cc')].sum()
    properties_3[2,filename,:] = net_c['throat.diameter'][net_c.throats('layer_Internal_cc')].sum()
    properties_3[3,filename,:] = net_c['throat.diameter'][net_c.throats('layer_Internal_mem')].sum()
    properties_3[4,filename,:] = net_c['throat.diameter'][net_c.throats('layer_mem')].sum()
    properties_2[5,filename,:] = net_c['pore.diameter'][net_c.pores('layer_cc')].sum()
    properties_2[6,filename,:] = net_c['pore.diameter'][net_c.pores('layer_Internal_cc')].sum()
    properties_2[7,filename,:] = net_c['pore.diameter'][net_c.pores('layer_Internal_mem')].sum()
    properties_2[8,filename,:] = net_c['pore.diameter'][net_c.pores('layer_mem')].sum()
    
    op.io.VTK.save(network=net_c, filename=filepath+electrode_name[filename]+'_3D_cathode_parts', phases=net_c.project['catholyte'])
    op.io.VTK.save(network=net_a, filename=filepath+electrode_name[filename]+'_3D_anode_parts', phases=net_a.project['anolyte'])
    op.io.VTK.save(network=net_c, filename=filepath+electrode_name[filename]+'_2D_mem', phases=net_c.project['catholyte'])    
    
    # =============================================================================
    # SAVE DATA TO AN EXCEL FILE 
    # Data storage
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=2).value = 'Generation 0'
    ws.cell(row=1, column=3).value = 'Generation 3'
    ws.cell(row=1, column=3).value = 'Generation 999'
        
    ws.cell(row=2, column=1).value = 'Average pore diameter'
    for k in range(len(electrode_name)):
        ws.cell(row=2, column=2+k).value = properties[0,k]           
    ws.cell(row=3, column=1).value = 'Average throat diameter'
    for k in range(len(electrode_name)):
        ws.cell(row=3, column=2+k).value = properties[1,k]       
    ws.cell(row=5, column=1).value = 'Pore surface area current collector'
    for k in range(len(electrode_name)):
        ws.cell(row=5, column=2+k).value = properties[2,k]      
    ws.cell(row=6, column=1).value = 'Pore surface area internal current collector'
    for k in range(len(electrode_name)):
        ws.cell(row=6, column=2+k).value = properties[3,k]   
    ws.cell(row=7, column=1).value = 'Pore surface area internal membrane'
    for k in range(len(electrode_name)):
        ws.cell(row=7, column=2+k).value = properties[4,k]
    ws.cell(row=8, column=1).value = 'Pore surface area membrane'
    for k in range(len(electrode_name)):
        ws.cell(row=8, column=2+k).value = properties[5,k]  
    ws.cell(row=10, column=1).value = 'Throat average diameter cc'
    for k in range(len(electrode_name)):
        ws.cell(row=10, column=2+k).value = properties[6,k]
    ws.cell(row=11, column=1).value = 'Throat average diameter icc'
    for k in range(len(electrode_name)):
        ws.cell(row=11, column=2+k).value = properties[7,k]
    ws.cell(row=12, column=1).value = 'Throat average diameter imem'
    for k in range(len(electrode_name)):
        ws.cell(row=12, column=2+k).value = properties[8,k]
    ws.cell(row=13, column=1).value = 'Throat average diameter mem'
    for k in range(len(electrode_name)):
        ws.cell(row=13, column=2+k).value = properties[9,k]
    ws.cell(row=15, column=1).value = 'Pore average diameter cc'
    for k in range(len(electrode_name)):
        ws.cell(row=15, column=2+k).value = properties[10,k]
    ws.cell(row=16, column=1).value = 'Pore average diameter icc'
    for k in range(len(electrode_name)):
        ws.cell(row=16, column=2+k).value = properties[11,k]
    ws.cell(row=17, column=1).value = 'Pore average diameter imem'
    for k in range(len(electrode_name)):
        ws.cell(row=17, column=2+k).value = properties[12,k]
    ws.cell(row=18, column=1).value = 'Pore average diameter mem'
    for k in range(len(electrode_name)):
        ws.cell(row=18, column=2+k).value = properties[13,k]
    ws.cell(row=20, column=1).value = 'Pore surface area cc'
    for k in range(len(electrode_name)):
        for l in range(len(properties_2[1,1,:])):
            ws.cell(row=20+k, column=2+l).value = properties_2[1,k,l]
    ws.cell(row=20+len(electrode_name), column=1).value = 'Pore surface area icc'
    for k in range(len(electrode_name)):
        for l in range(len(properties_2[1,1,:])):
            ws.cell(row=20+len(electrode_name)+k, column=2+l).value = properties_2[2,k,l]
    ws.cell(row=20+len(electrode_name)*2, column=1).value = 'Pore surface area imem'
    for k in range(len(electrode_name)):
        for l in range(len(properties_2[1,1,:])):
            ws.cell(row=20+len(electrode_name)*2+k, column=2+l).value = properties_2[3,k,l]
    ws.cell(row=20+len(electrode_name)*3, column=1).value = 'Pore surface area mem'
    for k in range(len(electrode_name)):
        for l in range(len(properties_2[1,1,:])):
            ws.cell(row=20+len(electrode_name)*3+k, column=2+l).value = properties_2[4,k,l]
    ws.cell(row=20+len(electrode_name)*4+1, column=1).value = 'Throat diameter cc'
    for k in range(len(electrode_name)):
        for l in range(len(properties_3[1,1,:])):
            ws.cell(row=20+len(electrode_name)*4+1+k, column=2+l).value = properties_3[1,k,l]
    ws.cell(row=20+len(electrode_name)*5+1, column=1).value = 'Throat diameter icc'
    for k in range(len(electrode_name)):
        for l in range(len(properties_3[1,1,:])):
            ws.cell(row=20+len(electrode_name)*5+1+k, column=2+l).value = properties_3[2,k,l]
    ws.cell(row=20+len(electrode_name)*6+1, column=1).value = 'Throat diameter imem'
    for k in range(len(electrode_name)):
        for l in range(len(properties_3[1,1,:])):
            ws.cell(row=20+len(electrode_name)*6+1+k, column=2+l).value = properties_3[3,k,l]
    ws.cell(row=20+len(electrode_name)*7+1, column=1).value = 'Throat diameter mem'
    for k in range(len(electrode_name)):
        for l in range(len(properties_3[1,1,:])):
            ws.cell(row=20+len(electrode_name)*7+1+k, column=2+l).value = properties_3[4,k,l]
    ws.cell(row=20+len(electrode_name)*8+2, column=1).value = 'Pore diameter cc'
    for k in range(len(electrode_name)):
        for l in range(len(properties_2[1,1,:])):
            ws.cell(row=20+len(electrode_name)*8+2+k, column=2+l).value = properties_2[5,k,l]
    ws.cell(row=20+len(electrode_name)*9+2, column=1).value = 'Pore diameter icc'
    for k in range(len(electrode_name)):
        for l in range(len(properties_2[1,1,:])):
            ws.cell(row=20+len(electrode_name)*9+2+k, column=2+l).value = properties_2[6,k,l]
    ws.cell(row=20+len(electrode_name)*10+2, column=1).value = 'Pore diameter imem'
    for k in range(len(electrode_name)):
        for l in range(len(properties_2[1,1,:])):
            ws.cell(row=20+len(electrode_name)*10+2+k, column=2+l).value = properties_2[7,k,l]
    ws.cell(row=20+len(electrode_name)*11+2, column=1).value = 'Pore diameter mem'
    for k in range(len(electrode_name)):
        for l in range(len(properties_2[1,1,:])):
            ws.cell(row=20+len(electrode_name)*11+2+k, column=2+l).value = properties_2[8,k,l]
    
    wb.save(filename = '.\\output\\' + Type + electrode_name[filename] + '\\' + electrode_name[filename] + 'properties.xlsx')
    
    # Plot length-thickness hexagonal binning plots for the current field
    surface = net_c.find_neighbor_throats(pores=net_c.pores('surface'))
    op.topotools.trim(network=net_c, throats=surface)
    op.topotools.trim(network=net_c, pores=net_c.pores('surface'))
    
    op.io.VTK.save(network=net_c, phases=catholyte, filename='.\\output\\' + Type + electrode_name[filename] + '\\' +'cathode_' + 'nosurface')  # Varying on your OS this output can cause an error, due to the decimal in E_cell
    op.utils.Workspace().save_project(project=project_1, filename='.\\output\\' + Type + electrode_name[filename] + '\\' + electrode_name[filename])

    project2 = op.io.PNM.load_project(filename='.\\output\\' + Type + electrode_name[filename] + '\\' + electrode_name[filename])
    test2 = project2.network
    project3 = op.io.PNM.load_project(filename='.\\output\\' + Type + electrode_name[filename] + '\\' + electrode_name[filename])
    test3 = project3.network    
    project4 = op.io.PNM.load_project(filename='.\\output\\' + Type + electrode_name[filename] + '\\' + electrode_name[filename])
    test4 = project4.network
    project5 = op.io.PNM.load_project(filename='.\\output\\' + Type + electrode_name[filename] + '\\' + electrode_name[filename])
    test5 = project5.network
    
    # op.topotools.trim(network=test2, throats=test2.throats('layer_cc'))
    # op.topotools.trim(network=test2, throats=test2.throats('layer_Internal_cc'))
    # op.topotools.trim(network=test2, throats=test2.throats('layer_Internal_mem'))
    # op.topotools.trim(network=test2, pores=test2.pores('layer_cc'))
    # op.topotools.trim(network=test2, pores=test2.pores('layer_Internal_cc'))
    # op.topotools.trim(network=test2, pores=test2.pores('layer_Internal_mem'))
    
    # op.io.VTK.save(network=test2, phases=catholyte, filename='.\\output\\' + Type + electrode_name[filename] + '\\' + 'cathode_' + 'mem')  # Varying on your OS this output can cause an error, due to the decimal in E_cell
    
    # op.topotools.trim(network=test3, throats=test3.throats('layer_cc'))
    # op.topotools.trim(network=test3, throats=test3.throats('layer_Internal_cc'))
    # op.topotools.trim(network=test3, throats=test3.throats('layer_mem'))
    # op.topotools.trim(network=test3, pores=test3.pores('layer_cc'))
    # op.topotools.trim(network=test3, pores=test3.pores('layer_Internal_cc'))
    # op.topotools.trim(network=test3, pores=test3.pores('layer_mem'))
    
    # op.io.VTK.save(network=test3, phases=catholyte, filename='.\\output\\' + Type + electrode_name[filename] + '\\cathode_' + 'imem')  # Varying on your OS this output can cause an error, due to the decimal in E_cell
    
    # op.topotools.trim(network=test4, throats=test4.throats('layer_cc'))
    # op.topotools.trim(network=test4, throats=test4.throats('layer_Internal_mem'))
    # op.topotools.trim(network=test4, throats=test4.throats('layer_mem'))
    # op.topotools.trim(network=test4, pores=test4.pores('layer_cc'))
    # op.topotools.trim(network=test4, pores=test4.pores('layer_Internal_mem'))
    # op.topotools.trim(network=test4, pores=test4.pores('layer_mem'))
    
    # op.io.VTK.save(network=test4, phases=catholyte, filename='.\\output\\' + Type + electrode_name[filename] + '\\cathode_' + 'icc')  # Varying on your OS this output can cause an error, due to the decimal in E_cell
    
    # op.topotools.trim(network=test5, throats=test5.throats('layer_mem'))
    # op.topotools.trim(network=test5, throats=test5.throats('layer_Internal_cc'))
    # op.topotools.trim(network=test5, throats=test5.throats('layer_Internal_mem'))
    # op.topotools.trim(network=test5, pores=test5.pores('layer_mem'))
    # op.topotools.trim(network=test5, pores=test5.pores('layer_Internal_cc'))
    # op.topotools.trim(network=test5, pores=test5.pores('layer_Internal_mem'))
    
    # op.io.VTK.save(network=test5, phases=catholyte, filename='.\\output\\' + Type + electrode_name[filename] + '\\cathode_' + 'cc')  # Varying on your OS this output can cause an error, due to the decimal in E_cell


# =================================================================================
# Save the absolute velocity to be able to plot the absolute velocity in the throats in paraview:     
Neigbor_array = net_c['throat.conns']
Pore_0 = net_c['throat.conns'][:,0]
Pore_1 = net_c['throat.conns'][:,1]
Pres_pore_0 = catholyte['pore.pressure'][net_c['throat.conns'][:,0]]
Pres_pore_1 = catholyte['pore.pressure'][net_c['throat.conns'][:,1]]
Delta_p = np.abs(Pres_pore_0 - Pres_pore_1)
g_h = catholyte['throat.hydraulic_conductance']
t_a = net_c['throat.area']
catholyte['throat.absolute_flowrate'] = Delta_p*g_h
catholyte['throat.absolute_velocity'] = catholyte['throat.absolute_flowrate'] / t_a

surface = net_c.find_neighbor_throats(pores=net_c.pores('surface'))
op.topotools.trim(network=net_c, throats=surface)
op.topotools.trim(network=net_c, pores=net_c.pores('surface'))

op.io.VTK.save(network=net_c, phases=catholyte, filename='.\\Genetic_Algorithm\\' + electrode_name[filename] + '_velocity')

# =================================================================================
# Save the concentration field and pore size distributions:
pore_volume=net_c['pore.volume'][net_c.pores('internal')]
pore_diameter = catholyte['pore.concentration'][net_c.pores('internal')]
pores=net_c.pores('internal')
bin_width = 1
pore_diameter_2 = pore_diameter * 1e2 
norm_pore_volume = pore_volume/np.sum(pore_volume) 
norm_pore_volume_2 = pore_volume/np.sum(pore_volume) 

# Set up iterating variables
max_pore_size = int(max(np.ceil(pore_diameter_2/bin_width)*bin_width))
pore_size_range = range(0, max_pore_size, bin_width)
norm_volume = np.zeros(len(pore_size_range)) 
norm_volume_2 = np.zeros(len(pore_size_range)) 
bins = np.zeros(len(pore_size_range))
i = 0

# Divide pores over bins
for size_bin in pore_size_range:
    nr = 0
    for pore in pores:
        if size_bin < pore_diameter_2[nr] <= size_bin + bin_width:
            
            bins[i] += 1
            norm_volume[i] += 1 #norm_pore_volume[nr]
            norm_volume_2[i] += norm_pore_volume_2[nr]
        nr = nr+1
    i += 1
    
# Test if all pores are assigned    
print(f'Total normalized volume (should be 1): {sum(norm_volume) :.2f}')
# Compute cumulative volume
j = 0
cum_volume_iter = 0
cum_volume = np.zeros(len(pore_size_range))
for bin_volume in norm_volume_2:
    cum_volume[j] = cum_volume_iter + bin_volume
    cum_volume_iter += bin_volume
    j += 1

psd = {}
psd['pore_size_range'] = pore_size_range
psd['bins'] = bins
psd['norm_volume'] = norm_volume
psd['cum_volume'] = cum_volume

plt.plot(pore_size_range, psd['norm_volume'])
plt.plot(pore_size_range, psd['cum_volume'])

# Output to excel
wb = openpyxl.Workbook()
ws = wb.active
output_variables = [psd['pore_size_range'], psd['norm_volume'], psd['cum_volume']]
output_names = ['pore size (PSD)', 'normalized volume', 'cumulative volume']
units = ['um', '-', '-']
for column_num in range(0, len(output_variables)): # column_num represents the variables in ouput_variables.
    ws.cell(row=1, column=column_num+1).value = output_names[column_num]
    ws.cell(row=2, column=column_num+1).value = units[column_num]
    ws.cell(row=3, column=column_num+1).value = None
    ws.cell(row=4, column=column_num+1).value = None
    for row_num in range(0, len(output_variables[column_num])): # row_num represents the ith entry within the variable array
        ws.cell(row = row_num + 1, column = column_num + 1).value = output_variables[column_num][row_num] # row_num + 5 to convert to Origin format 
wb.save(filename = '.\\output\\' + electrode_name[filename] + '_concentration.xlsx')